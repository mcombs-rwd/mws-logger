# bacnet_controllers.py
import logging

from openpyxl import load_workbook

from bacnet_table import BacnetController

logger = logging.getLogger(__name__)
logger.info("Logging initialized.")

def _parse_obj_cell(cell_value):
    """Parse cell into obj_type and obj_id
    
    <Cell 'Sheet1'.C2> <analog-value,     126>
    """
    (obj_type, obj_id) = cell_value[1:-1].split(',')
    if obj_type.strip().lower() in ['analog-value', 'analog value', 'analogvalue']:
        obj_type = 'analogValue'
    elif obj_type.strip().lower() in ['multi-state-value', 'multistatevalue']:
        obj_type = 'multiStateValue'
    return (obj_type, int(obj_id))

def parse_spreadsheet(spreadsheet, WORKBOOK_NUM=0):
    """Reads spreadsheet, returns a list of (obj_type, obj_id, obj_name) tuples
    
    ARGUMENTS:
        spreadsheet: path to an xlxs file
        WORKBOOK_NUM: position of workbook (tab) to read
        NAME_COL: column number of names column
        OBJ_COL: column with objects
    
        row: A list of cells:
        <Cell 'Sheet1'.B2> VAV-5-401_AirFl
        <Cell 'Sheet1'.C2> <analog-value,     126>
    """
    # todo: change name_col and obj_col to letters
    # columns are numbered from 0–n
    NAME_COL = 1  
    OBJ_COL = 2    
    try:
        workbook = load_workbook(filename=spreadsheet)
        logger.debug(f"...Loading {spreadsheet}: {workbook.sheetnames}")
    except:
        logger.warning(f"Error: File {spreadsheet} doesn't exist")
        return []
    # todo: accept workbook name instead of number
    current_sheet = workbook[workbook.sheetnames[WORKBOOK_NUM]]  # Use 1st sheet
    objects = []
    for (n, row) in enumerate(current_sheet.iter_rows(min_row=2, max_col=3)):
        # breakpoint()
        if row[NAME_COL].value == None:
            continue
        (obj_type, obj_id) = _parse_obj_cell(row[OBJ_COL].value)
        mo = (obj_type, obj_id, row[NAME_COL].value)
        logger.debug(mo)
        objects.append(mo)
    return objects

def load_controllers(automation_servers):
    '''Create list of devices to log and attributes from each.

    Arguments:
        automation_servers: list of ip addrs and spreadsheets with object IDs
    
    Returns:
        A list of Controllers.
    '''
    # Parse the name column into the room/device and attribute.
    controllers = []
    for (n, server) in enumerate(automation_servers):
        if not ('ip' in server and 'file' in server):
            logger.warning(f"Error: Server {n} missing ip or file attribute.")
            continue
        monitored_objects = parse_spreadsheet(server['file'])
        co = BacnetController(server['ip'], monitored_objects, server['desc'])
        logger.debug(co)
        controllers.append(co)
    return controllers
